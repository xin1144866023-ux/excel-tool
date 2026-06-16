from __future__ import annotations

from pathlib import Path
import json
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


WHITE_SIDE = Side(style="thin", color="FFFFFF")
DEFAULT_BORDER = Border(left=WHITE_SIDE, right=WHITE_SIDE, top=WHITE_SIDE, bottom=WHITE_SIDE)


def css_rgb_to_hex(value: str | None) -> str | None:
    if not value or value == "transparent" or value.startswith("rgba"):
        return None
    match = re.search(r"rgb\((\d+),\s*(\d+),\s*(\d+)\)", value)
    if not match:
        return None
    return "".join(f"{int(part):02X}" for part in match.groups())


def px_to_points(value: str | None, default: float = 10) -> float:
    if not value:
        return default
    match = re.search(r"([\d.]+)px", value)
    if not match:
        return default
    return max(8, round(float(match.group(1)) * 0.75, 1))


def as_cell_value(text: str | None):
    if text is None:
        return None
    text = text.strip()
    if not text:
        return None
    cleaned = text.replace(",", "")
    if re.fullmatch(r"\d+", cleaned):
        return int(cleaned)
    if re.fullmatch(r"\d+\.\d+", cleaned):
        return float(cleaned)
    return text


def apply_style(ws, min_row: int, max_row: int, min_col: int, max_col: int, style: dict) -> None:
    fill_color = css_rgb_to_hex(style.get("backgroundColor"))
    font_color = css_rgb_to_hex(style.get("color")) or "000000"
    font_weight = style.get("fontWeight", "400")
    bold = str(font_weight).isdigit() and int(font_weight) >= 600
    font_size = px_to_points(style.get("fontSize"), default=10)
    text_align = style.get("textAlign") or "center"
    horizontal = "left" if text_align in {"left", "start"} else "right" if text_align in {"right", "end"} else "center"

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=fill_color) if fill_color else PatternFill(fill_type=None)
            cell.font = Font(name="Arial", size=font_size, bold=bold, color=font_color)
            cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
            cell.border = DEFAULT_BORDER


def place_table_cells(rows: list[dict], start_row: int) -> tuple[list[dict], list[float]]:
    occupied: set[tuple[int, int]] = set()
    placements: list[dict] = []
    widths: list[float] = []

    for row_data in rows:
        excel_row = start_row + int(row_data["rowIndex"])
        excel_col = 1
        for cell_data in row_data["cells"]:
            while (excel_row, excel_col) in occupied:
                excel_col += 1

            row_span = max(int(cell_data.get("rowSpan") or 1), 1)
            col_span = max(int(cell_data.get("colSpan") or 1), 1)
            min_row = excel_row
            max_row = excel_row + row_span - 1
            min_col = excel_col
            max_col = excel_col + col_span - 1

            while len(widths) < max_col:
                widths.append(0)
            per_col_width = float(cell_data.get("rect", {}).get("width", 72)) / col_span
            for col_idx in range(min_col - 1, max_col):
                widths[col_idx] = max(widths[col_idx], per_col_width)

            placements.append(
                {
                    "min_row": min_row,
                    "max_row": max_row,
                    "min_col": min_col,
                    "max_col": max_col,
                    "cell": cell_data,
                }
            )

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    occupied.add((row, col))
            excel_col = max_col + 1

    return placements, widths


def write_title(ws, meta: dict, max_cols: int) -> None:
    title_items = meta.get("titleItems") or []
    title = title_items[0]["text"] if title_items else "Life Insurance"
    profile = title_items[1]["text"] if len(title_items) > 1 else ""
    split_col = max(1, max_cols - 2)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=split_col)
    ws.merge_cells(start_row=1, start_column=split_col + 1, end_row=1, end_column=max_cols)
    for col in range(1, max_cols + 1):
        cell = ws.cell(1, col)
        cell.fill = PatternFill("solid", fgColor="D0CCCC")
        cell.border = DEFAULT_BORDER
        cell.font = Font(name="Arial", size=15, bold=True, color="000000")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.cell(1, 1, title)
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(1, split_col + 1, profile)
    ws.cell(1, split_col + 1).font = Font(name="Arial", size=10, bold=True, color="000000")
    ws.cell(1, split_col + 1).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 8


def write_table(ws, table: dict, start_row: int = 3) -> int:
    rows = table["rows"]
    placements, widths = place_table_cells(rows, start_row)
    max_cols = max(len(widths), 1)

    for idx, pixel_width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(45, max(8, round(pixel_width / 7.5, 1)))

    for row_data in rows:
        excel_row = start_row + int(row_data["rowIndex"])
        row_height = float(row_data.get("rect", {}).get("height", 20))
        ws.row_dimensions[excel_row].height = max(5, round(row_height * 0.75, 1))

    for placement in placements:
        min_row = placement["min_row"]
        max_row = placement["max_row"]
        min_col = placement["min_col"]
        max_col = placement["max_col"]
        cell_data = placement["cell"]

        if max_row > min_row or max_col > min_col:
            ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

        apply_style(ws, min_row, max_row, min_col, max_col, cell_data.get("style", {}))
        top_left = ws.cell(row=min_row, column=min_col, value=as_cell_value(cell_data.get("text")))
        if isinstance(top_left.value, int):
            top_left.number_format = "#,##0"
        elif isinstance(top_left.value, float):
            top_left.number_format = "#,##0.00"

    return start_row + len(rows)


def write_notes(ws, meta: dict, start_row: int, max_cols: int) -> None:
    note_items = meta.get("noteItems") or []
    if not note_items:
        return

    row = start_row + 1
    for index, item in enumerate(note_items):
        text = item.get("text", "")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
        cell = ws.cell(row=row, column=1, value=text)
        if index == 0:
            cell.fill = PatternFill("solid", fgColor=css_rgb_to_hex(item.get("style", {}).get("backgroundColor")) or "B85E17")
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 18
        else:
            cell.font = Font(name="Arial", size=10, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = max(18, min(64, 14 + len(text) / 7))
        for col in range(1, max_cols + 1):
            ws.cell(row=row, column=col).border = Border()
        row += 1


def build_workbook(payload: dict, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "HTML 鏈接"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"

    placements, widths = place_table_cells(payload["table"]["rows"], 3)
    max_cols = max(len(widths), 1)
    write_title(ws, payload.get("meta", {}), max_cols)
    next_row = write_table(ws, payload["table"], start_row=3)
    write_notes(ws, payload.get("meta", {}), next_row, max_cols)

    ws.print_area = f"A1:{get_column_letter(max_cols)}{ws.max_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def verify_workbook(output_path: Path) -> None:
    wb = load_workbook(output_path)
    if wb.sheetnames != ["HTML 鏈接"]:
        raise RuntimeError(f"unexpected sheets: {wb.sheetnames}")
    if any(len(wb[s]._images) for s in wb.sheetnames):
        raise RuntimeError("output should not contain embedded images")
    ws = wb["HTML 鏈接"]
    if not ws["A1"].value:
        raise RuntimeError("missing page title")
    if ws.max_row < 10 or ws.max_column < 3:
        raise RuntimeError("output table is unexpectedly small")


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: converter.py input_payload.json output.xlsx", file=sys.stderr)
        return 2
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    build_workbook(payload, output_path)
    verify_workbook(output_path)
    print(f"wrote {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
